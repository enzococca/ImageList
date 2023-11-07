import os

import openpyxl
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import filedialog, simpledialog
import time
import tkinter.messagebox
from PIL import Image, ImageTk, ImageOps, UnidentifiedImageError
import piexif
from pathlib2 import Path
import shutil
from tkinterdnd2 import DND_FILES, TkinterDnD
import tempfile
from geopandas import GeoDataFrame
from shapely.geometry import Point
import folium
from folium.plugins import AntPath
import math
import pandas as pd
import webbrowser
import imgkit
from openpyxl.drawing.image import Image as XLImage
import json

image_label = None
global listbox


def show_error_in_listbox(error_message, listbox):
    # print(error_message)
    listbox.insert(tk.END, f"ERROR: {error_message}")


def show_in_listbox(message, listbox):
    # print(error_message)
    listbox.insert(tk.END, f"{message}")


def load_processed_files(json_path):
    try:
        with open(json_path, "r") as f:
            return set(json.load(f))
    except FileNotFoundError:
        return set()


def save_processed_files(json_path, processed_files_set):
    with open(json_path, "w") as f:
        json.dump(list(processed_files_set), f)


def create_widgets(root):
    try:
        global image_label

        # Create a frame for the treeview
        tree_frame = ttk.Frame(root, padding="10")
        tree_frame.grid(row=0, column=0, sticky="nsew")

        # Create the treeview inside the frame
        tree = ttk.Treeview(tree_frame, padding="10")
        tree.grid(row=0, column=0, sticky="nsew")

        # Create a frame for the image label

        image_frame = ttk.Frame(root, padding="10")
        image_frame.grid(row=1, column=1, sticky="nsew", padx=10, pady=10)

        # Create the image label inside the frame
        image_label = tk.Label(image_frame)
        image_label.pack()
    except Exception as e:
        show_error_in_listbox(f"Error in create_widgets: {e}", listbox)


def on_item_double_click(tree, event):
    try:
        item = tree.selection()[0]
        filepath = tree.item(item, "values")[0]

        # Crea una nuova finestra Toplevel
        new_window = tk.Toplevel()
        new_window.title("Image Preview")

        canvas = tk.Canvas(new_window, width=800, height=600)
        canvas.pack(fill="both", expand="yes")
        new_window.grid_rowconfigure(0, weight=1)
        new_window.grid_columnconfigure(0, weight=1)

        original_image = Image.open(filepath)
        zoom_factor = 1.0
        x, y = 0, 0

        def resize_image(event=None, zoom=1.0):
            nonlocal zoom_factor
            zoom_factor *= zoom
            width, height = original_image.size
            new_size = int(width * zoom_factor), int(height * zoom_factor)
            image = original_image.resize(new_size, Image.LANCZOS)
            tk_image = ImageTk.PhotoImage(image)
            canvas.create_image(10, 10, anchor="ne", image=tk_image)
            canvas.image = tk_image  # Mantenere un riferimento all'immagine
            canvas.config(scrollregion=canvas.bbox("all"))

        def zoom_image(event):
            if event.delta > 0:
                resize_image(zoom=1.1)
            elif event.delta < 0:
                resize_image(zoom=0.9)

        def on_button_press(event):
            nonlocal x, y
            x, y = event.x, event.y

        def on_button_release(event):
            nonlocal x, y
            x, y = None, None

        def on_move_press(event):
            nonlocal x, y
            dx = event.x - x
            dy = event.y - y
            canvas.move("all", dx, dy)
            x, y = event.x, event.y

        resize_image()

        canvas.bind("<MouseWheel>", zoom_image)
        canvas.bind("<ButtonPress-1>", on_button_press)
        canvas.bind("<ButtonRelease-1>", on_button_release)
        canvas.bind("<B1-Motion>", on_move_press)

    except Exception as e:
        print(f"Errore nel doppio clic: {e}")


# def show_image_preview(tree, listbox):
#     try:
#         global image_label
#         selected_items = tree.selection()  # Get the ID of the selected item
#         if selected_items:  # Check if there is a selected item
#             selected_item_id = selected_items[0]
#             file_path = tree.set(
#                 selected_item_id, "fullpath"
#             )  # Get the full path of the selected item
#
#             _, ext = os.path.splitext(file_path)
#             if ext.lower() in [".jpeg", ".jpg"]:
#                 # Open and resize the image
#                 image = Image.open(file_path)
#                 image = image.resize(
#                     (200, 200), Image.LANCZOS
#                 )  # Resize the image to 100x100 pixels
#
#                 # Create a PhotoImage object and set it as the image option of the image label
#                 photo_image = ImageTk.PhotoImage(image)
#                 image_label.config(image=photo_image)
#
#                 # Store the PhotoImage object as an attribute of the label to prevent it from being garbage collected
#                 image_label.photo_image = photo_image
#     except UnidentifiedImageError:
#         show_error_in_listbox("Invalid image format.", listbox)
#     except Exception as e:
#         show_error_in_listbox(f"Error in show_image_preview: {e}", listbox)

# Inizializzazione
magnifier_window = None
original_image = None  # Sarà impostato nel metodo show_image_preview


def show_magnifier(event):
    global magnifier_window, magnifier_label, original_image

    if original_image:  # Verifica se original_image è stato impostato
        x, y = event.x, event.y
        cropped_image = original_image.crop((x - 25, y - 25, x + 25, y + 25))
        cropped_image = cropped_image.resize((300, 300), Image.LANCZOS)
        tk_cropped_image = ImageTk.PhotoImage(cropped_image)

        if magnifier_window:
            magnifier_label.config(image=tk_cropped_image)
            magnifier_label.image = tk_cropped_image
            magnifier_window.geometry(f"+{event.x_root+10}+{event.y_root+10}")
        else:
            magnifier_window = tk.Toplevel()
            magnifier_label = tk.Label(magnifier_window, image=tk_cropped_image)
            magnifier_label.image = tk_cropped_image
            magnifier_label.pack()
            magnifier_window.geometry(f"+{event.x_root+10}+{event.y_root+10}")


def hide_magnifier(event):
    global magnifier_window
    if magnifier_window:
        magnifier_window.destroy()
        magnifier_window = None


def show_image_preview(root, tree, listbox):
    try:
        global original_image  # ora è una variabile globale
        selected_items = tree.selection()
        print(f"Selected items: {selected_items}")  # Debugging
        if selected_items:
            selected_item_id = selected_items[0]
            file_path = tree.set(selected_item_id, "fullpath")
            print(f"File path: {file_path}")  # Debugging
            _, ext = os.path.splitext(file_path)

            if ext.lower() in [
                ".jpeg",
                ".jpg",
                ".png",
            ]:  # Aggiunto ".png" come formato accettabile
                original_image = Image.open(file_path)  # Qui impostiamo original_image
                original_image = original_image.resize((200, 200), Image.LANCZOS)
                photo_image = ImageTk.PhotoImage(original_image)
                image_label.config(image=photo_image)
                image_label.photo_image = photo_image
                # Creazione del canvas
                canvas = tk.Canvas(root, width=200, height=200)
                canvas.grid(row=1, column=1)  # posiziona il canvas dove preferisci
                canvas.create_image(0, 0, anchor=tk.NW, image=photo_image)
                canvas.grid()

                canvas.bind("<Motion>", show_magnifier)
                canvas.bind("<Leave>", hide_magnifier)

                print("Image should be displayed now.")  # Debugging

    except UnidentifiedImageError:
        print("Invalid image format.")  # Debugging
    except Exception as e:
        print(f"Error in show_image_preview: {e}")  # Debugging


def populate_tree(tree, node):
    try:
        path = Path(tree.set(node, "fullpath"))
        if path.is_dir():
            for p in path.iterdir():
                if p.suffix.lower() != ".ini":  # Ignore .ini files
                    is_dir = p.is_dir()
                    oid = tree.insert(
                        node,
                        "end",
                        text=p.name,
                        values=[p, "directory" if is_dir else "file"],
                    )
                    if is_dir:
                        tree.after(10, populate_tree, tree, oid)
    except Exception as e:
        show_error_in_listbox(f"Error in populate_tree: {e}", listbox)
    else:
        show_in_listbox(f"populate tree: Ok!", listbox)


def get_exif_data(image_path):
    try:
        exif_dict = piexif.load(image_path)
        return exif_dict
    except Exception as e:
        show_error_in_listbox(f"Failed to get EXIF data: {e}", listbox)


def get_coordinates(gps_info):
    try:
        lat_data = gps_info[piexif.GPSIFD.GPSLatitude]
        lon_data = gps_info[piexif.GPSIFD.GPSLongitude]

        # Convert the GPS coordinates stored in the EXIF to dd format
        lat_degree = lat_data[0][0] / lat_data[0][1]
        lat_minute = lat_data[1][0] / lat_data[1][1]
        lat_second = lat_data[2][0] / lat_data[2][1]
        lon_degree = lon_data[0][0] / lon_data[0][1]
        lon_minute = lon_data[1][0] / lon_data[1][1]
        lon_second = lon_data[2][0] / lon_data[2][1]

        latitude = lat_degree + (lat_minute / 60) + (lat_second / 3600)
        longitude = lon_degree + (lon_minute / 60) + (lon_second / 3600)

        # Adjust the sign of the DD values based on the hemisphere
        if gps_info[piexif.GPSIFD.GPSLatitudeRef] == b"S":
            latitude = -latitude
        if gps_info[piexif.GPSIFD.GPSLongitudeRef] == b"W":
            longitude = -longitude

        return latitude, longitude
    except Exception as e:
        show_error_in_listbox(f"Failed to get coordinates data: {e}", listbox)


def degrees_to_direction(degrees):
    try:
        directions = ["N", "NE", "E", "SE", "S", "SW", "W", "NW", "N"]
        index = round(degrees / 45)
        return directions[index]
    except Exception as e:
        show_error_in_listbox(f"Failed to get degree direction data: {e}", listbox)


output_geojson_file = None
output_shapefile_dir = None


def process_images(
    directory,
    output_file,
    progress_var,
    total_files,
    root,
    time_label,
    file_count_label,
    listbox,
    progress_label,
    json_path,
):

    row = 2  # Sposta questa riga all'inizio della funzione
    img_direction = None  # Initialize to None or some default value
    compass_direction_str = None  # Initialize to None or some default value
    processed_files_set = load_processed_files(json_path)

    image_data_list = []
    # Set to keep track of processed files
    duplicate_files_count = 0  # Counter for duplicate files
    print("il file excel è:" f"{ output_file}")
    if output_file is None:
        output_file = tempfile.mktemp(suffix=".xlsx")
    # If the Excel file already exists, read it and add the names of the images to the set
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
        for row in ws.iter_rows(
            min_row=2, max_col=1, values_only=True
        ):  # Skip the header row
            if row[0] is not None:
                processed_files_set.add(
                    os.path.join(directory, row[0])
                )  # Add the full path to the set
        #wb.close()  # Close the workbook after reading the file names

    else:
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "Filename",
                "Latitude",
                "Longitude",
                "DateTime",
                "Orientation",
                "OrientationDegrees",
                "Folder",
            ]
        )
        row = 2

    for cell in ws["1:1"]:
        cell.font = Font(bold=True)
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = 20
    print(f"Total files: {total_files}")
    start_time = time.time()
    processed_files = 0
    for dirpath, dirs, files in os.walk(directory):
        for filename in files:
            if filename.lower().endswith((".jpg", ".jpeg")):
                full_path = os.path.join(dirpath, filename)
                exif_data = get_exif_data(full_path)
                try:
                    datetime_ = exif_data["Exif"][
                        piexif.ExifIFD.DateTimeOriginal
                    ].decode("utf-8")
                except Exception as e:
                    show_error_in_listbox(f"Error getting DATETIME info: {e}", listbox)
                    datetime_ = "N/A"
                # Create a unique identifier for each file based on its name and DateTime
                unique_file_identifier = f"{full_path}"
                if (
                    unique_file_identifier in processed_files_set
                ):  # If the file has already been processed
                    duplicate_files_count += (
                        1  # Increment the counter for duplicate files
                    )
                    continue  # Skip this file
                processed_files_set.add(
                    unique_file_identifier
                )  # Add the unique identifier to the set

                # After processing a file, write the processed files set to a file
                # with open('processed_files.json', 'w') as f:
                # json.dump(list(processed_files_set), f)

                try:
                    folder_path = Path(full_path).parent
                    relative_folder_path = folder_path.relative_to(directory)
                    if (
                        folder_path.name.startswith("F")
                        or folder_path.name.startswith("ART")
                        or folder_path.name.startswith("OVERVIEW")
                    ):
                        folder_name = str(relative_folder_path).replace(os.sep, "-")
                    else:
                        folder_name = f"{folder_path.name}"
                    exif_data = get_exif_data(full_path)
                    # Print the entire EXIF dictionary in a sorted order
                    for ifd in ("0th", "Exif", "GPS", "1st"):
                        for tag in sorted(exif_data[ifd]):
                            tag_name = piexif.TAGS[ifd][tag]["name"]
                            value = exif_data[ifd][tag]
                            # print(f"{tag_name}: {value}")

                    # Get the image direction in degrees
                    if piexif.GPSIFD.GPSImgDirection in exif_data["GPS"]:
                        img_direction = exif_data["GPS"][piexif.GPSIFD.GPSImgDirection]
                        img_direction = (
                            img_direction[0] / img_direction[1]
                        )  # Get the actual value
                        # Convert the image direction to a compass direction
                        compass_direction = degrees_to_direction(img_direction)
                        compass_direction_str = (
                            compass_direction
                            + "("
                            + str(img_direction)
                            + "\u00b0"
                            + ")"
                        )

                    else:
                        compass_direction = "N/A"
                        # compass_direction_str = 'N/A'
                    # Get the GPS info
                    try:
                        gps_info = exif_data["GPS"]
                        if gps_info:
                            latitude, longitude = get_coordinates(gps_info)
                        else:
                            latitude = "N/A"
                            longitude = "N/A"
                    except Exception as e:
                        show_error_in_listbox(f"Error getting GPS info: {e}", listbox)
                        latitude = "N/A"
                        longitude = "N/A"
                    # Assicurati che 'Exif' sia un dizionario e che contenga la chiave 'DateTimeOriginal'
                    try:
                        with Image.open(full_path) as image:
                            # Perform your operations on the image
                            thumbnail = ImageOps.exif_transpose(image)
                            thumbnail.thumbnail((100, 100))
                            thumbnail_dir = os.path.join(directory, "thumbnail")
                            os.makedirs(thumbnail_dir, exist_ok=True)

                            thumbnail_path = os.path.join(
                                thumbnail_dir,
                                f"{relative_folder_path}_{filename}_thumbnail.jpg",
                            )
                            # Extract the directory from the thumbnail path
                            thumbnail_dir_path = os.path.dirname(thumbnail_path)

                            # Create the directory if it doesn't exist
                            os.makedirs(thumbnail_dir_path, exist_ok=True)

                            # Now you can safely save the thumbnail
                            thumbnail.save(thumbnail_path)

                            image_data_list.append(
                                {
                                    "Filename": filename,
                                    "Latitude": latitude,
                                    "Longitude": longitude,
                                    "DateTime": datetime_,
                                    "Orientation": compass_direction_str,  # + '(' + str(img_direction) + '\u00b0' + ')',
                                    "OrientationDegrees": img_direction,  # exif_data['GPS'][piexif.GPSIFD.GPSImgDirection][0] / exif_data['GPS'][piexif.GPSIFD.GPSImgDirection][1],
                                    "Folder": folder_name,
                                }
                            )
                            if "DateTime" in exif_data["Exif"]:
                                print("DateTime:", exif_data["Exif"]["DateTime"])
                            else:
                                print(
                                    # "DateTime not found in EXIF data.",
                                    exif_data["Exif"],
                                )

                            try:
                                # Debug: stampa i dati che stiamo per aggiungere
                                print(
                                    f"Adding row :{row.index}, {filename}, {latitude}, {longitude}, {datetime_}, {compass_direction_str}, {img_direction}, {folder_name}"
                                )

                                ws.append(
                                    [
                                        filename,
                                        latitude,
                                        longitude,
                                        datetime_,
                                        compass_direction_str,
                                        img_direction,
                                        folder_name,
                                    ]
                                )
                                # # Verifica che data_to_append non sia una tupla singola
                                if (
                                    isinstance(data_to_append, tuple)
                                    and len(data_to_append) == 1
                                ):
                                    data_to_append = list(data_to_append[0])
                                #
                                ws.append(data_to_append)
                                # Debugging: Verify that the file exists and is accessible
                                if not os.path.exists(thumbnail_path):
                                    print(f"File does not exist: {thumbnail_path}")
                                elif not os.access(thumbnail_path, os.R_OK):
                                    print(
                                        f"File is not accessible (read permissions): {thumbnail_path}"
                                    )
                                else:
                                    print(
                                        f"File is present and accessible: {thumbnail_path}"
                                    )
                                    img = XLImage(thumbnail_path)
                                    img.width = img.width * 1.5
                                    img.height = img.height * 1.5
                                    ws.add_image(img, f"H{row}")
                                    ws.column_dimensions["H"].width = img.width // 7
                                    ws.row_dimensions[row].height = img.height

                                    row += 1
                                    processed_files += 1  # Increment the counter for the processed files
                                    elapsed_time = time.time() - start_time
                                    remaining_files = total_files - processed_files
                                    remaining_time = (
                                        elapsed_time / processed_files * remaining_files
                                    )
                                    time_label[
                                        "text"
                                    ] = f"Estimated time remaining: {remaining_time:.2f} seconds"
                                    file_count_label[
                                        "text"
                                    ] = f"Files remaining: {remaining_files}"
                                    listbox.insert(
                                        tk.END,
                                        f"Processed file {filename} - {latitude}, {longitude} - {compass_direction}",
                                    )
                                    root.update()

                                    progress_var.set(
                                        processed_files
                                    )  # Update the progress variable after the image has been processed
                                    progress_percent = (
                                        processed_files / total_files
                                    ) * 100  # Calculate the progress percentage
                                    progress_label[
                                        "text"
                                    ] = f"{progress_percent:.2f}%"  # Update the progress label
                            except (IOError, OSError) as e:
                                show_error_in_listbox(
                                    f"Error processing image {filename}: {e}", listbox
                                )
                                continue  # Skip this file and move on to the next one

                            except Exception as e:
                                show_error_in_listbox(
                                    f"Error processing file {filename}: {e}", listbox
                                )

                    except (IOError, OSError) as e:
                        # Handle the error, possibly log it to the listbox or console
                        show_error_in_listbox(
                            f"Error processing image {filename}: {e}", listbox
                        )
                        continue  # Skip this file and move on to the next one

                except Exception as e:
                    show_error_in_listbox(
                        f"Error processing file {filename}: {e}", listbox
                    )


    geo_data = [
        d for d in image_data_list if d["Latitude"] != "N/A" and d["Longitude"] != "N/A"
    ]

    # Crea un GeoDataFrame
    geometry = [
        Point(xy)
        for xy in zip(
            [float(d["Longitude"]) for d in geo_data],
            [float(d["Latitude"]) for d in geo_data],
        )
    ]
    geo_df = GeoDataFrame(geo_data, geometry=geometry)

    # Salva in GeoJSON
    if output_geojson_file:  # Se output_geojson_file è stato impostato
        geo_df.to_file(output_geojson_file, driver="GeoJSON")

    # Salva in shapefile
    if output_shapefile_dir:  # Se output_shapefile_dir è stato impostato
        geo_df.to_file(output_shapefile_dir, driver="ESRI Shapefile")


    save_processed_files(json_path, processed_files_set)

    tkinter.messagebox.showinfo(
        "Information",
        f"Finished processing images. {duplicate_files_count} duplicate images were not added.",
    )

    try:
        wb.save(output_file)
    except Exception as e:
        print(f"Exception occurred while saving the workbook: {e}")
        import traceback

        traceback.print_exc()  # Stampa il traceback completo per l'eccezione

    thumbnail_dir = os.path.join(directory, "thumbnail")
    if os.path.exists(thumbnail_dir):
        shutil.rmtree(thumbnail_dir)  # Remove the thumbnail directory
    # Once all files are processed, save the list of processed files


input_dir = None
output_file = None


def add_map_to_excel(ws, lista_coordinate, img_path="ciao.jpg"):
    config = imgkit.config(
        wkhtmltoimage="C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltoimage.exe"
    )

    # Calcola latitudine e longitudine medie per centrare la mappa
    lat_medio = sum(lat for lat, _ in lista_coordinate) / len(lista_coordinate)
    lon_medio = sum(lon for _, lon in lista_coordinate) / len(lista_coordinate)
    # Crea una mappa Folium
    ma = folium.Map(location=[lat_medio, lon_medio], zoom_start=9)
    # Aggiungi i marker dalla lista di coordinate
    for lat, lon in lista_coordinate:
        folium.Marker([lat, lon]).add_to(ma)

    html_path = "map.html"
    ma.save(html_path)

    # Opzioni per imgkit
    options = {
        "width": "1024",
        "height": "768",
        "crop-w": "500",
        "crop-h": "500",
        "crop-x": "50",
        "crop-y": "50",
        "quality": "100",
        "log-level": "info",
    }
    # Converti HTML in PNG usando imgkit
    imgkit.from_file(html_path, img_path, config=config, options=options)
    # Inserisci l'immagine in Excel
    img = XLImage(img_path)
    img.width = 640  # Imposta la larghezza
    img.height = 480  # Imposta l'altezza
    cell_width = img.width // 2  # Adjust 'some_factor' as needed
    cell_height = img.height // 2  # Adjust 'some_factor' as needed
    ws.column_dimensions["J"].width = cell_width
    ws.row_dimensions[2].height = cell_height
    ws.add_image(img, "J2")

    # Regola le dimensioni delle celle per farle corrispondere all'immagine
    # ws.column_dimensions['J'].width = 24
    # ws.row_dimensions[2].height = 80


def drop(event, tree, root):  # Add root as an argument
    print("Drop event triggered.")  # Print a message when the drop event is triggered
    global input_dir
    files = root.tk.splitlist(
        event.data
    )  # Use splitlist() to get the list of dropped files
    for file in files:
        _, ext = os.path.splitext(file)
        if ext.lower() in [".jpeg", ".jpg"]:
            # print(f"Source file: {file}")

            # Add the file to the tree
            tree.insert("", "end", text=file, values=[file, "file"])

            # Copy the file to the selected directory in the tree
            selected_items = tree.selection()  # Get the ID of the selected item
            if selected_items:  # Check if there is a selected item
                selected_item_id = selected_items[0]
                input_dir = tree.set(
                    selected_item_id, "fullpath"
                )  # Get the full path of the selected item

                if os.path.isdir(input_dir):  # Make sure the destination is a directory
                    # print(f"Destination directory: {input_dir}")
                    shutil.copy(file, input_dir)
                    root.update_idletasks()  # Update the UI
                    # Update the tree
                    tree.delete(
                        *tree.get_children(selected_item_id)
                    )  # Delete the current children of the selected item
                    populate_tree(
                        tree, selected_item_id
                    )  # Repopulate the tree with the new directory structure

                else:
                    show_error_in_listbox(f"{input_dir} is not a directory.", listbox)
            else:
                show_error_in_listbox("No directory selected in the tree.", listbox)


base_dir = None  # Define base_dir at the start of your program


def create_directories(tree):
    global input_dir, base_dir  # Dichiarazione di input_dir come variabile globale

    base_dir = filedialog.askdirectory()  # Ask the user to choose a base directory
    prefix = simpledialog.askstring(
        "Input", "Enter the prefix for directories:"
    )  # Ask the user to input a prefix

    input_dir = base_dir  # Imposta input_dir sul base_dir

    range_str = simpledialog.askstring(
        "Input", "Enter the range of directories to create (e.g., '1-10'):"
    )
    ranges = range_str.split(";")
    for range_str in ranges:
        if "-" in range_str:
            start, end = map(int, range_str.split("-"))

            for i in range(start, end + 1):
                dir_name = "{}{}".format(
                    prefix, i
                )  # This will create names like N00001, N00002, etc.
                dir_path = os.path.join(base_dir, dir_name)
                os.mkdir(dir_path)
        else:
            # Gestione di una singola directory
            dir_name = "{}{}".format(prefix, range_str)
            dir_path = os.path.join(base_dir, dir_name)
            os.mkdir(dir_path)
    # Clear the tree before populating it
    for i in tree.get_children():
        tree.delete(i)

    # Populate the tree with the new directory structure
    root_node = tree.insert("", "end", text=base_dir, values=[base_dir, "directory"])
    populate_tree(tree, root_node)
    if input_dir and output_file:  # Se sia input_dir che output_file sono impostati
        start_button["state"] = "normal"  # Abilita il pulsante di avvio


def add_subdirectories(tree):
    global input_dir  # Access the global variable base_dir
    if input_dir is None:  # Check if base_dir has been defined
        messagebox.showerror(
            "Error", "Please create directories before adding subdirectories."
        )
        return
    dir_path = filedialog.askdirectory()  # Ask the user to choose a directory
    sub_dirs = simpledialog.askstring(
        "Input", "Enter the names of subdirectories to create, separated by semicolons:"
    )
    sub_dirs = [
        x.strip() for x in sub_dirs.split(";")
    ]  # Split the string into a list of subdirectory names

    # Create subdirectories in the selected directory
    for sub_dir in sub_dirs:
        sub_dir_path = os.path.join(dir_path, sub_dir)
        os.mkdir(sub_dir_path)

    # Clear the tree before populating it
    for i in tree.get_children():
        tree.delete(i)

    # Populate the tree with the new directory structure
    root_node = tree.insert("", "end", text=input_dir, values=[input_dir, "directory"])
    populate_tree(tree, root_node)


def import_images(
    progress_var, time_label, file_count_label, listbox, root, progress_bar, tree
):
    try:
        global input_dir  # Dichiarazione di input_dir come variabile globale
        input_dir = filedialog.askdirectory(
            title="Seleziona la cartella delle immagini"
        )
        total_files = sum(
            filename.lower().endswith((".jpg", ".jpeg"))
            for dirpath, dirs, files in os.walk(input_dir)
            for filename in files
        )
        show_in_listbox(f"{total_files}", listbox)
        # Ottieni il numero totale di file immagine
        progress_var.set(0)  # Inizializza la variabile di progresso
        progress_bar[
            "maximum"
        ] = total_files  # Imposta il valore massimo della barra di avanzamento
        if input_dir and output_file:  # Se sia input_dir che output_file sono impostati
            start_button["state"] = "normal"  # Abilita il pulsante di avvio

        # Pulisci l'albero prima di popolarlo
        for i in tree.get_children():
            tree.delete(i)

        # Popola l'albero con la struttura della directory
        root_node = tree.insert(
            "", "end", text=input_dir, values=[input_dir, "directory"]
        )
        populate_tree(tree, root_node)
    except Exception as e:
        show_error_in_listbox(f"Failed to get import image data: {e}", listbox)


def save_excel():
    try:
        global output_file  # Declare output_file as a global variable
        output_file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Excel File As",
        )
        if input_dir and output_file:  # If both input_dir and output_file are set
            start_button["state"] = "normal"  # Enable the start button
    except Exception as e:
        show_error_in_listbox(f"Failed to get excel data: {e}", listbox)


def save_geojson():
    try:
        global output_geojson_file  # Declare output_geojson_file as a global variable
        output_geojson_file = filedialog.asksaveasfilename(
            defaultextension=".geojson",
            filetypes=[("GeoJSON Files", "*.geojson")],
            title="Save GeoJSON File As",
        )
        if (
            input_dir and output_geojson_file
        ):  # If both input_dir and output_geojson_file are set
            start_button["state"] = "normal"  # Enable the start button
    except Exception as e:
        show_error_in_listbox(f"Failed to get GeoJSON data: {e}", listbox)


def save_shapefile():
    try:
        global output_shapefile_dir  # Declare output_shapefile_dir as a global variable
        output_shapefile_dir = filedialog.askdirectory(title="Save Shapefile Folder As")
        if (
            input_dir and output_shapefile_dir
        ):  # If both input_dir and output_shapefile_dir are set
            start_button["state"] = "normal"  # Enable the start button
    except Exception as e:
        show_error_in_listbox(f"Failed to get shapefile data: {e}", listbox)


def get_json_path(output_file):
    directory, excel_file = os.path.split(output_file)
    json_file = os.path.splitext(output_file)[0] + ".json"
    return os.path.join(directory, json_file)


def start_processing(
    progress_var, time_label, file_count_label, listbox, root, progress_label
):
    total_files = sum(
        filename.lower().endswith((".jpg", ".jpeg"))
        for _, _, files in os.walk(input_dir)
        for filename in files
    )  # Get the total number of JPEG images

    json_path = get_json_path(output_file)
    progress_var.set(0)  # Initialize the progress variable
    process_images(
        input_dir,
        output_file,
        progress_var,
        total_files,
        root,
        time_label,
        file_count_label,
        listbox,
        progress_label,
        json_path,
    )


def autoscroll(sbar, first, last):
    """Hide and show scrollbar as needed."""
    first, last = float(first), float(last)
    if first <= 0 and last >= 1:
        sbar.grid_remove()
    else:
        sbar.grid()
    sbar.set(first, last)


selected_item = None  # Variabile globale per tenere traccia dell'elemento selezionato


def on_item_select(event):
    global selected_item
    tree = event.widget
    selected_items = tree.selection()
    if not selected_items:  # Se la selezione è vuota
        return  # Esci dalla funzione
    selected_item = selected_items[0]
    print(f"Item selected: {selected_item}")


def on_item_drop(event, widget=None):
    global selected_item
    src_path = ""
    dest_dir = ""
    tree = event.widget
    target_item = tree.identify("item", event.x, event.y)
    if target_item and selected_item:
        if selected_item and tree.exists(selected_item):
            src_path = tree.item(selected_item)["values"][0]
            dest_dir = tree.item(target_item)["values"][0]

        # Verifica se src_path è un'immagine
        # if src_path and src_path.lower().endswith(('.jpg', '.jpeg')):
        if src_path.lower().endswith((".jpg", ".jpeg")):
            if os.path.isdir(
                dest_dir
            ):  # Assicurati che la destinazione sia una directory
                try:
                    # Esegui lo spostamento o la copia qui
                    shutil.move(
                        src_path, os.path.join(dest_dir, os.path.basename(src_path))
                    )
                    print(f"Spostato {src_path} in {dest_dir}")

                    # Trova il nodo padre per aggiornare solo quella parte dell'albero
                    parent_item = tree.parent(target_item)
                    update_tree(tree, parent_item)

                except Exception as e:
                    print(f"Errore durante lo spostamento: {e}")
            else:
                print(f"{dest_dir} non è una directory.")
        else:
            print("Il file trascinato non è un'immagine supportata.")


def get_expanded_nodes(tree, node):
    expanded_nodes = []
    for child in tree.get_children(node):
        if tree.item(child, "open"):
            expanded_nodes.append(tree.set(child, "fullpath"))
            expanded_nodes.extend(get_expanded_nodes(tree, child))
    return expanded_nodes


def expand_nodes(tree, node, paths_to_expand):
    for child in tree.get_children(node):
        child_path = tree.set(child, "fullpath")
        if child_path in paths_to_expand:
            tree.item(child, open=True)
            expand_nodes(tree, child, paths_to_expand)


def update_tree(tree, parent_item):
    # Ottiene la lista degli elementi espansi
    expanded_nodes = get_expanded_nodes(tree, parent_item)

    # Elimina tutti i nodi figli
    tree.delete(*tree.get_children(parent_item))

    # Reinserisce i nodi
    populate_tree(tree, parent_item)

    # Riespande i nodi
    expand_nodes(tree, parent_item, expanded_nodes)


def main():
    global input_dir, output_file, start_button, listbox  # Declare input_dir, output_file, and start_button as global variables

    root = TkinterDnD.Tk()  # Define root2 here
    root.drop_target_register(DND_FILES)
    root.dnd_bind(
        "<<Drop>>", lambda event: drop(event, tree, root)
    )  # Remove root2 from the lambda function
    # Preliminare definizione della Treeview senza yscrollcommand
    tree = ttk.Treeview(root, columns=("fullpath", "type"), displaycolumns="")

    # Definizione della scrollbar verticale
    vsb = ttk.Scrollbar(root, orient="vertical", command=tree.yview)

    # Aggiornamento della Treeview per utilizzare la scrollbar
    tree.configure(yscrollcommand=vsb.set)
    hsb = ttk.Scrollbar(orient="horizontal")

    # tree = ttk.Treeview(columns=("fullpath", "type"), displaycolumns="")

    create_widgets(root)

    tree.bind(
        "<<TreeviewSelect>>", lambda event: show_image_preview(root, tree, listbox)
    )
    tree.bind("<ButtonRelease-1>", lambda event, widget=None: on_item_drop(event))
    tree.bind(
        "<Button-1>", on_item_select
    )  # Aggiungi questo bind per gestire la pressione del pulsante del mouse
    tree.bind("<Double-1>", lambda event: on_item_double_click(tree, event))
    hsb["command"] = tree.xview

    tree.heading("#0", text="Directory Structure", anchor="w")
    # Configurazione della Treeview per utilizzare la scrollbar

    tree.grid(column=0, row=1, sticky="nsew", padx=10, pady=10)
    vsb.grid(column=1, row=1, sticky="ns", pady=10)
    hsb.grid(column=0, row=2, sticky="ew", padx=10)
    # hsb.grid(column=0, row=2, sticky="ew")

    root.grid_columnconfigure(0, weight=1)
    root.grid_rowconfigure(1, weight=1)  # Change this to 1
    root.title("Image Processor")

    menubar = tk.Menu(root)
    filemenu = tk.Menu(menubar, tearoff=0)
    filemenu.add_command(
        label="Import photos folder",
        command=lambda: import_images(
            progress_var,
            time_label,
            file_count_label,
            listbox,
            root,
            progress_bar,
            tree,
        ),
    )

    export_vector_menu = tk.Menu(filemenu, tearoff=0)
    export_vector_menu.add_command(label="Save GeoJSON", command=save_geojson)
    export_vector_menu.add_command(label="Save Shapefile", command=save_shapefile)

    filemenu.add_cascade(
        label="Vector Export", menu=export_vector_menu
    )  # Aggiunge il sottomenu al menu File
    filemenu.add_command(label="Save Excel", command=save_excel)
    filemenu.add_command(
        label="Create Directories", command=lambda: create_directories(tree)
    )

    menubar.add_cascade(label="File", menu=filemenu)
    root.config(menu=menubar)

    frame = tk.Frame(root)
    frame.grid(row=0, column=0, sticky="nsew")

    listbox = tk.Listbox(frame)  # Create a listbox to show the process in real time
    listbox.grid(row=0, column=0, sticky="nsew")

    scrollbar = tk.Scrollbar(frame, orient="vertical", command=listbox.yview)
    scrollbar.grid(row=0, column=1, sticky="ns")

    listbox.configure(yscrollcommand=scrollbar.set)
    progress_var = tk.DoubleVar()

    progress_frame = tk.Frame(root)  # Create a frame to hold the progress bar and label
    progress_frame.grid(row=3, column=0, sticky="ew")

    progress_bar = ttk.Progressbar(progress_frame, length=600, variable=progress_var)
    progress_bar.pack(fill="x")

    progress_label = tk.Label(
        progress_frame, text="0.00%"
    )  # Create a label to show the progress percentage
    progress_label.place(
        relx=0.5, rely=0.5, anchor="center"
    )  # Place the label at the center of the progress bar

    start_button = tk.Button(
        root,
        text="Start",
        state="disabled",
        command=lambda: start_processing(
            progress_var, time_label, file_count_label, listbox, root, progress_label
        ),
    )  # Create a label to show the progress percentage

    start_button.grid(row=2, column=0, sticky="ew")
    time_label = tk.Label(root)  # Create a label to show the estimated time remaining
    time_label.grid(row=3, column=0, sticky="w")
    file_count_label = tk.Label(root)  # Create a label to show the file count
    file_count_label.grid(row=5, column=0, sticky="w")

    # Add button for adding subdirectories

    add_button = tk.Button(
        root, text="Add subdirectory", command=lambda: add_subdirectories(tree)
    )
    add_button.grid(row=5, column=0, sticky="n")

    root.grid_rowconfigure(1, weight=1)
    root.grid_columnconfigure(0, weight=1)
    frame.grid_rowconfigure(1, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    root.mainloop()


if __name__ == "__main__":
    main()
